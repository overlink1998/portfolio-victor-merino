import tkinter as tk
from tkinter import messagebox, filedialog, simpledialog
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os
from datetime import datetime
import calendar
import locale
import threading 
import pyodbc
import tkinter as tk
from tkinter import messagebox
import sys


try:
    locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')
except Exception as e:
    print("No fue posible establecer la configuración regional: " + str(e))
def clean_and_fill(df):
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            df[col].fillna(0, inplace=True)
        else:
            df[col].fillna('NA', inplace=True)
    return df




def threaded_function(func, *args):
    thread = threading.Thread(target=func, args=args)
    thread.start()
    return thread

def format_dataframe_dates(df):
    """ Formatea las columnas de fecha de un DataFrame para quitar la hora. """
    date_cols = [col for col in df.columns if 'Fecha' in col or 'FECHA' in col]  # Busca columnas que contienen fechas
    for col in date_cols:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime('%d/%m/%Y')  # Cambia formato a 'YYYY-MM-DD'
    return df


def update_email_in_prorrogas(file_path, email):
    """Actualiza la columna de correo en la hoja de prórrogas con un email específico."""
    wb = load_workbook(file_path)
    if 'Hoja3' in wb.sheetnames:
        ws = wb['Hoja3']
        email_column = 9  # Asumiendo que la columna de correo es la novena columna
        for row in range(2, ws.max_row + 1):  # Comienza en 2 para omitir el encabezado
            ws.cell(row=row, column=email_column).value = email
        wb.save(file_path)
        print("Todos los correos en 'Hoja3' han sido actualizados con éxito.")
    else:
        print("La hoja 'Hoja3' no existe en el archivo destino.")


def save_individual_reports(start_date, end_date, db_path, password):
    if not db_path or not password:
        messagebox.showerror("Error", "Database path or password not provided.")
        return

    conexion_str = f"DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={db_path};PWD={password}"
    try:
        conn = pyodbc.connect(conexion_str)
        df_contratos = pd.read_sql("SELECT * FROM CONTRATOS", conn)
        df_prorrogas = pd.read_sql("SELECT * FROM PRORROGAS", conn)
        df_ventas = pd.read_sql("SELECT * FROM VENTAS", conn)
        conn.close()

        if getattr(sys, 'frozen', False):
            # Si es empaquetado, usar la ruta del ejecutable
            ruta_documentos = os.path.dirname(sys.executable)
        else:
            # De lo contrario, usar el directorio del script
            ruta_documentos = os.path.dirname(os.path.abspath(__file__))

        ruta_excel = os.path.join(ruta_documentos, "Informe.xlsx")
        with pd.ExcelWriter(ruta_excel, engine='openpyxl') as writer:
            df_contratos.to_excel(writer, sheet_name='Contratos', index=False)
            df_prorrogas.to_excel(writer, sheet_name='Prorrogas', index=False)
            df_ventas.to_excel(writer, sheet_name='Ventas', index=False)

        print("Informe generado correctamente en:", ruta_excel)

    except Exception as e:
        messagebox.showerror("Error", f"Failed to connect or execute query: {str(e)}")
        return

    try:
        fecha_inicio = pd.to_datetime(start_date, dayfirst=True)
        fecha_fin = pd.to_datetime(end_date, dayfirst=True)
        if fecha_inicio > fecha_fin:
            messagebox.showerror("Error", "La fecha de inicio debe ser antes de la fecha de fin.")
            return
        nombre_mes_inicio = fecha_inicio.strftime('%B').capitalize()
        nombre_mes_fin = fecha_fin.strftime('%B').capitalize()
        if fecha_inicio.month == fecha_fin.month:
            nombre_fecha = f"{fecha_inicio.day}-{fecha_fin.day} de {nombre_mes_inicio} {fecha_inicio.year}"
        else:
            nombre_fecha = f"{fecha_inicio.day} de {nombre_mes_inicio} - {fecha_fin.day} de {nombre_mes_fin} {fecha_fin.year}"

        # Formatear el nombre de la carpeta según mes y añog
        #nombre_mes = calendar.month_name[fecha_inicio.month]
        #anio = fecha_inicio.year
        folder_name = f"{nombre_fecha.lower()}"
        folder_path = os.path.join(ruta_documentos, folder_name)
        os.makedirs(folder_path, exist_ok=True)

        ruta_excel = os.path.join(ruta_documentos, "Informe.xlsx")
        xls = pd.ExcelFile(ruta_excel)
        df_informe = pd.read_excel(xls, sheet_name=0)
        df_prorrogas = pd.read_excel(xls, sheet_name=1)
        df_ventas = pd.read_excel(xls, sheet_name=2)

        df_informe = clean_and_fill(df_informe)
        df_prorrogas = clean_and_fill(df_prorrogas)
        df_ventas = clean_and_fill(df_ventas)
        

        # Convertir fechas
        df_informe['Fecha  Contrato'] = pd.to_datetime(df_informe['Fecha  Contrato'], errors='coerce', dayfirst=True)
        df_informe['Fecha Retirado'] = pd.to_datetime(df_informe['Fecha Retirado'], errors='coerce', dayfirst=True)
        df_prorrogas['Fecha'] = pd.to_datetime(df_prorrogas['Fecha'], errors='coerce', dayfirst=True)
        df_ventas['FECHA VENTA'] = pd.to_datetime(df_ventas['FECHA VENTA'], errors='coerce', dayfirst=True)

        # Filtrar datos
        df_contratos = df_informe[(df_informe['Fecha  Contrato'] >= fecha_inicio) & (df_informe['Fecha  Contrato'] <= fecha_fin)]
        df_prorrogas = df_prorrogas[(df_prorrogas['Fecha'] >= fecha_inicio) & (df_prorrogas['Fecha'] <= fecha_fin)]
        df_retiros = df_informe[(df_informe['Fecha Retirado'] >= fecha_inicio) & (df_informe['Fecha Retirado'] <= fecha_fin) & (df_informe['Retirado'] == True)]
        df_sacas = df_informe[(df_informe['Fecha Retirado'] >= fecha_inicio) & (df_informe['Fecha Retirado'] <= fecha_fin) & (df_informe['Saca'] == True)]
        df_ventas = df_ventas[(df_ventas['FECHA VENTA'] >= fecha_inicio) & (df_ventas['FECHA VENTA'] <= fecha_fin)]
        df_contratos_no_retirados = df_contratos[df_contratos['Fecha Retirado'].isna() | pd.to_datetime(df_contratos['Fecha Retirado'], errors='coerce').isna()]        


        df_retiros_combinados = pd.concat([df_retiros, df_sacas]).drop_duplicates(subset=['No Contrato']).reset_index(drop=True)
        df_prorrogas_final = df_prorrogas.merge(df_contratos[['No Contrato', 'Nombre Cliente', 'Cedula', 'Telefono', 'dirección', 'Correo']], on='No Contrato', how='left')
        df_prorrogas_final = df_prorrogas_final[['No Contrato', 'Fecha', 'Numero Meses', 'Valor Pagado', 'Nombre Cliente', 'Cedula', 'Telefono', 'dirección', 'Correo']]
        df_sacas = df_sacas[['No Contrato', 'Fecha  Contrato', 'Descripcion', 'Peso', 'Valor', 'Nombre Cliente', 'Cedula', 'Telefono', 'dirección', 'Correo', 'Sprecio', 'Fecha Retirado' ]]
        df_retiros = df_retiros[['No Contrato', 'Fecha  Contrato', 'Descripcion', 'Peso', 'Valor', 'Nombre Cliente', 'Cedula', 'Telefono', 'dirección', 'Correo', 'Sprecio', 'Fecha Retirado' ]]
        df_contratos_no_retirados = df_contratos_no_retirados[['No Contrato', 'Fecha  Contrato', 'Descripcion', 'Peso', 'Valor', 'Nombre Cliente', 'Cedula', 'Telefono' , 'Campo1','Correo', 'Fecha Retirado']]


        df_contratos = df_contratos[['No Contrato', 'Fecha  Contrato', 'Descripcion', 'Peso', 'Valor', 'Nombre Cliente', 'Cedula', 'Telefono', 'dirección', 'Correo']]
        df_ventas['Nombre Cliente'] = 'NA'
        df_ventas['Cedula'] = 'NA'
        df_ventas['Telefono'] = 'NA'
        df_ventas['dirección'] = 'NA'
        df_ventas['Correo'] = 'NA'
        df_ventas = df_ventas[['CODIGO', 'COSTO VENTA', 'UTILIDAD', 'FECHA VENTA']]
        df_prorrogas_final = clean_and_fill(df_prorrogas_final) 
        
        df_contratos.sort_values('Fecha  Contrato', inplace=True)
        df_contratos.sort_values('No Contrato', inplace=True)
        df_retiros.sort_values('Fecha Retirado', inplace=True)
        df_sacas.sort_values('Fecha Retirado', inplace=True)
        df_prorrogas_final.sort_values('Fecha', inplace=True)
        df_ventas.sort_values('FECHA VENTA', inplace=True)
        df_contratos_no_retirados.sort_values('Fecha  Contrato', inplace=True)
        df_contratos_no_retirados.sort_values('No Contrato', inplace=True)
        

        

        #nombre_mes = calendar.month_name[fecha_inicio.month]
        #anio = fecha_inicio.year
        #nombre_fecha = f"{nombre_01mes.lower()} {anio}"
        df_contratos = format_dataframe_dates(df_contratos)
        df_prorrogas_final = format_dataframe_dates(df_prorrogas_final)
        df_retiros = format_dataframe_dates(df_retiros)
        df_sacas = format_dataframe_dates(df_sacas)        
        df_ventas = format_dataframe_dates(df_ventas)
        df_contratos_no_retirados = format_dataframe_dates(df_contratos_no_retirados)
        
        
        
        for sheet_name, df in zip(['Contratos', 'Prórrogas', 'Retiros', 'Ventas', 'contratos_no_retirados', 'df_sacas'], [df_contratos, df_prorrogas_final, df_retiros, df_ventas, df_contratos_no_retirados, df_sacas]):
            file_name = f"{sheet_name}_{nombre_fecha}.xlsx"
            file_path = os.path.join(folder_path, file_name)
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
            

        ruta_documentos = ruta_documentos
        nombre_archivo_origen = "Informe.xlsx"
        ruta_documentos2 = folder_path
        nombre_archivo_destino = f"Prórrogas_{nombre_fecha}.xlsx"

        ruta_archivo_origen = os.path.join(ruta_documentos, nombre_archivo_origen)
        df_origen = pd.read_excel(ruta_archivo_origen)

        ruta_archivo_destino = os.path.join(ruta_documentos2, nombre_archivo_destino)
        book = load_workbook(ruta_archivo_destino)


        if 'Prórrogas' not in book.sheetnames:
            print("La Prórrogas no existe en el archivo destino.")
        else:
           
            hoja3 = book['Prórrogas']
            max_row = hoja3.max_row

            for index, row in df_origen.iterrows():
                
                for r in range(1, max_row + 1):
                    cell_value = hoja3.cell(row=r, column=1).value  
                    if cell_value == row['No Contrato']:
                        
                        hoja3.cell(row=r, column=5).value = row['Nombre Cliente']
                        hoja3.cell(row=r, column=6).value = row['Cedula']
                        hoja3.cell(row=r, column=7).value = row['Telefono']
                        hoja3.cell(row=r, column=8).value = row['dirección']
                        hoja3.cell(row=r, column=9).value = row['Correo']
                        
                        break  

            
            book.save(ruta_archivo_destino)
            print("La Prórrogas del archivo ha sido actualizada con éxito.")


        ruta_documentos = ruta_documentos
        nombre_archivo_origen = "Informe.xlsx"
        ruta_documentos2 = folder_path

        nombre_archivo_destino = f"Prórrogas_{nombre_fecha}.xlsx"
        ruta_archivo_origen = os.path.join(ruta_documentos, nombre_archivo_origen)
        ruta_archivo_destino = os.path.join(ruta_documentos2, nombre_archivo_destino)
        df_origen = pd.read_excel(ruta_archivo_origen)


        def update_email_in_prorrogas(file_path, email):
            """Actualiza la columna de correo en la hoja de prórrogas con un email específico."""
            wb = load_workbook(file_path)
            if 'Prórrogas' in wb.sheetnames:
                ws = wb['Prórrogas']
                email_column = 9  # Asumiendo que la columna de correo es la novena columna
                for row in range(2, ws.max_row + 1):  # Comienza en 2 para omitir el encabezado
                    ws.cell(row=row, column=email_column).value = email
                wb.save(file_path)
                print("Todos los correos en 'Prórrogas' han sido actualizados con éxito.")
            else:
                print("La hoja 'Prórrogas' no existe en el archivo destino.")

        # Rutas y nombres de archivo

        ruta_documentos = ruta_documentos2

        nombre_archivo_destino = f"Prórrogas_{nombre_fecha}.xlsx"

        ruta_archivo_destino = os.path.join(ruta_documentos, nombre_archivo_destino)

        # Llamar a la función para actualizar correos
        update_email_in_prorrogas(ruta_archivo_destino, "cvlaemeralda1997@gmail.com")
        messagebox.showinfo("Éxito", f"Archivos guardados correctamente en la carpeta: {folder_path}")
    except Exception as e:
        messagebox.showerror("Error", str(e))
        
        
        
        
def generate_report(start_date, end_date, text_output, db_path, password):
    global conmas, nombre_mes, folder_name
    if not db_path:
        messagebox.showerror("Error", "Database path not provided.")
        return

    if password:
        conexion_str = f"DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={db_path};PWD={password}"
    else:
        conexion_str = f"DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={db_path}"
    try:
        conn = pyodbc.connect(conexion_str)
        df_contratos = pd.read_sql("SELECT * FROM CONTRATOS", conn)
        df_prorrogas = pd.read_sql("SELECT * FROM PRORROGAS", conn)
        df_ventas = pd.read_sql("SELECT * FROM VENTAS", conn)
        conn.close()
        if getattr(sys, 'frozen', False):
            # Si es empaquetado, usar la ruta del ejecutable
            ruta_documentos = os.path.dirname(sys.executable)
        else:
            # De lo contrario, usar el directorio del script
            ruta_documentos = os.path.dirname(os.path.abspath(__file__))


        ruta_excel = os.path.join(ruta_documentos, "Informe.xlsx")
        with pd.ExcelWriter(ruta_excel, engine='openpyxl') as writer:
            df_contratos.to_excel(writer, sheet_name='Contratos', index=False)
            df_prorrogas.to_excel(writer, sheet_name='Prorrogas', index=False)
            df_ventas.to_excel(writer, sheet_name='Ventas', index=False)

        print("Informe generado correctamente en:", ruta_excel)

    except Exception as e:
        messagebox.showerror("Error", f"Failed to connect or execute query: {str(e)}")
        return
    text_output.delete('1.0', tk.END)  # Clear existing text
    try:
        fecha_inicio = pd.to_datetime(start_date, dayfirst=True)
        fecha_fin = pd.to_datetime(end_date, dayfirst=True)

        if fecha_inicio > fecha_fin:
            text_output.insert(tk.END, "Error: La fecha de inicio debe ser antes que la fecha de fin.\n")
            return
        
        ruta_excel = os.path.join(ruta_documentos, "Informe.xlsx")
        xls = pd.ExcelFile(ruta_excel)
        df_informe = pd.read_excel(xls, sheet_name=0)
        df_prorrogas = pd.read_excel(xls, sheet_name=1)
        df_ventas = pd.read_excel(xls, sheet_name=2)

        df_informe = clean_and_fill(df_informe)
        df_prorrogas = clean_and_fill(df_prorrogas)
        df_ventas = clean_and_fill(df_ventas)

        # Convertir fechas
        df_informe['Fecha  Contrato'] = pd.to_datetime(df_informe['Fecha  Contrato'], errors='coerce', dayfirst=True)
        df_informe['Fecha Retirado'] = pd.to_datetime(df_informe['Fecha Retirado'], errors='coerce', dayfirst=True)
        df_prorrogas['Fecha'] = pd.to_datetime(df_prorrogas['Fecha'], errors='coerce', dayfirst=True)
        df_ventas['FECHA VENTA'] = pd.to_datetime(df_ventas['FECHA VENTA'], errors='coerce', dayfirst=True)

        # Filtrar datos
        df_contratos = df_informe[(df_informe['Fecha  Contrato'] >= fecha_inicio) & (df_informe['Fecha  Contrato'] <= fecha_fin)]
        df_prorrogas = df_prorrogas[(df_prorrogas['Fecha'] >= fecha_inicio) & (df_prorrogas['Fecha'] <= fecha_fin)]
        df_retiros = df_informe[(df_informe['Fecha Retirado'] >= fecha_inicio) & (df_informe['Fecha Retirado'] <= fecha_fin) & (df_informe['Retirado'] == True)]
        df_sacas = df_informe[(df_informe['Fecha Retirado'] >= fecha_inicio) & (df_informe['Fecha Retirado'] <= fecha_fin) & (df_informe['Saca'] == True)]
        df_ventas = df_ventas[(df_ventas['FECHA VENTA'] >= fecha_inicio) & (df_ventas['FECHA VENTA'] <= fecha_fin)]

        df_retiros_combinados = pd.concat([df_retiros, df_sacas]).drop_duplicates(subset=['No Contrato']).reset_index(drop=True)
        df_prorrogas_final = df_prorrogas.merge(df_contratos[['No Contrato', 'Nombre Cliente', 'Cedula', 'Telefono', 'dirección', 'Correo']], on='No Contrato', how='left')
        df_prorrogas_final = df_prorrogas_final[['No Contrato', 'Fecha', 'Numero Meses', 'Valor Pagado', 'Nombre Cliente', 'Cedula', 'Telefono', 'dirección', 'Correo']]
        df_retiros_combinados = df_retiros_combinados[['No Contrato', 'Fecha  Contrato', 'Descripcion', 'Peso', 'Valor', 'Nombre Cliente', 'Cedula', 'Telefono', 'dirección', 'Correo', 'Sprecio', 'Fecha Retirado' ]]

        df_contratos = df_contratos[['No Contrato', 'Fecha  Contrato', 'Descripcion', 'Peso', 'Valor', 'Nombre Cliente', 'Cedula', 'Telefono', 'dirección', 'Correo']]
        df_ventas['Nombre Cliente'] = 'NA'
        df_ventas['Cedula'] = 'NA'
        df_ventas['Telefono'] = 'NA'
        df_ventas['dirección'] = 'NA'
        df_ventas['Correo'] = 'NA'
        df_ventas = df_ventas[['CODIGO', 'COSTO VENTA', 'UTILIDAD', 'FECHA VENTA', 'Nombre Cliente', 'Cedula', 'Telefono', 'dirección', 'Correo']]
        df_prorrogas_final = clean_and_fill(df_prorrogas_final) 
        
        df_contratos.sort_values('Fecha  Contrato', inplace=True)
        df_contratos.sort_values('No Contrato', inplace=True)
        df_retiros_combinados.sort_values('Fecha Retirado', inplace=True)
        df_prorrogas_final.sort_values('Fecha', inplace=True)
        df_ventas.sort_values('FECHA VENTA', inplace=True)
        nombre_mes = calendar.month_name[fecha_inicio.month]
        anio = fecha_inicio.year
        nombre_mes_inicio = calendar.month_name[fecha_inicio.month]
        nombre_mes_fin = calendar.month_name[fecha_fin.month]
        
        
        df_contratos = format_dataframe_dates(df_contratos)
        df_prorrogas_final = format_dataframe_dates(df_prorrogas_final)
        df_retiros_combinados = format_dataframe_dates(df_retiros_combinados)
        df_ventas = format_dataframe_dates(df_ventas)
        

        if fecha_inicio.month == fecha_fin.month:
            nombre_fecha = f"{fecha_inicio.day}-{fecha_fin.day} {nombre_mes_inicio.lower()} {anio}"
        else:
            nombre_fecha = f"{fecha_inicio.day} {nombre_mes_inicio.lower()} - {fecha_fin.day} {nombre_mes_fin.lower()} {anio}"
        ruta_salida_excel = os.path.join(ruta_documentos, f"Informe_{nombre_fecha}.xlsx")
        coversion = "Informe_"
        conmas = coversion + nombre_fecha
        
        
        with pd.ExcelWriter(ruta_salida_excel, engine='openpyxl') as writer:
            df_contratos.to_excel(writer, sheet_name='Hoja1', index=False)
            df_retiros_combinados.to_excel(writer, sheet_name='Hoja2', index=False)
            df_prorrogas_final.to_excel(writer, sheet_name='Hoja3', index=False)
            df_ventas.to_excel(writer, sheet_name='Hoja4', index=False)
        
        
        
        
        
        ruta_archivo_origen = ruta_excel
        df_origen = pd.read_excel(ruta_archivo_origen)
        df_origen = clean_and_fill(df_origen)  # Limpieza de datos
        
        # Cargar el libro de Excel destino y actualizar la hoja 'Hoja3'
        ruta_archivo_destino = ruta_salida_excel
        book = load_workbook(ruta_archivo_destino)
        
        if 'Hoja3' in book.sheetnames:
            hoja3 = book['Hoja3']
            max_row = hoja3.max_row
            for index, row in df_origen.iterrows():
                for r in range(2, max_row + 1):  # Comienza en 2 para omitir el encabezado
                    cell_value = hoja3.cell(row=r, column=1).value
                    if cell_value == row['No Contrato']:
                        hoja3.cell(row=r, column=5).value = row.get('Nombre Cliente', 'NA')
                        hoja3.cell(row=r, column=6).value = row.get('Cedula', 'NA')
                        hoja3.cell(row=r, column=7).value = row.get('Telefono', 'NA')
                        hoja3.cell(row=r, column=8).value = row.get('dirección', 'NA')
                        hoja3.cell(row=r, column=9).value = row.get('Correo', 'NA')
            book.save(ruta_archivo_destino)
        else:
            print("La hoja 'Hoja3' no existe en el archivo destino.")
        
        
        
        
        
 
        
        text_output.insert(tk.END, f"Informe generado correctamente para el periodo: {start_date} a {end_date}\n")
        text_output.insert(tk.END, "Guardado en: " + ruta_salida_excel + "\n")
    except Exception as e:
        text_output.insert(tk.END, "Error al generar el informe: " + str(e) + "\n")
        


def select_database_file(entry_widget):
    """ Allows user to select a database file and updates the entry widget. """
    file_path = filedialog.askopenfilename(title="Select Access Database",
                                           filetypes=[("Microsoft Access Database", "*.accdb;*.mdb")])
    if file_path:
        entry_widget.delete(0, tk.END)
        entry_widget.insert(0, file_path)

def on_closing():
    if messagebox.askokcancel("Salir", "¿Estás seguro de que quieres salir?"):
        root.destroy()


def create_gui():
    global root, label_color
    root = tk.Tk()
    root.title("Generador de Informes")
    root.configure(background='black')
    icon_path = 'icono.ico'  # Asegúrate de que la ruta sea correcta
    
    # Intenta establecer la imagen como ícono de la ventana
    try:
        root.iconbitmap(icon_path)
        
    except tk.TclError as e:
        print(f"No se pudo cargar el ícono de la aplicación: {e}")

    label_color = 'gray'
    input_bg = 'black'
    input_fg = 'white'
    button_bg = 'green'
    button_fg = 'black'


    tk.Label(root, text="Fecha de inicio (DD/MM/AAAA):", bg='black', fg=button_bg).grid(row=1, column=0)
    start_date_entry = tk.Entry(root, bg=input_bg, fg=input_fg)
    start_date_entry.grid(row=1, column=1)

    tk.Label(root, text="Fecha de fin (DD/MM/AAAA):", bg='black', fg=button_bg).grid(row=2, column=0)
    end_date_entry = tk.Entry(root, bg=input_bg, fg=input_fg)
    end_date_entry.grid(row=2, column=1)

    tk.Label(root, text="Ingrese dirección de BD:", bg='black', fg=button_bg).grid(row=3, column=0)
    db_path_entry = tk.Entry(root, bg=input_bg, fg=input_fg)
    db_path_entry.grid(row=3, column=1)
    browse_button = tk.Button(root, text="CARGAR RUTA", bg=button_bg, fg=button_fg, command=lambda: select_database_file(db_path_entry))
    browse_button.grid(row=3, column=2)

    tk.Label(root, text="Ingrese contraseña de BD:", bg='black', fg=button_bg).grid(row=4, column=0)
    password_entry = tk.Entry(root, show="*", bg=input_bg, fg=input_fg)
    password_entry.grid(row=4, column=1)

    text_output = tk.Text(root, height=10, width=50, bg=input_bg, fg=input_fg)
    text_output.grid(row=5, column=0, columnspan=3)

    generate_button = tk.Button(root, text="Generar Informe", bg=button_bg, fg=button_fg,
                                command=lambda: threaded_function(generate_report,
                                                                  start_date_entry.get(),
                                                                  end_date_entry.get(),
                                                                  text_output,
                                                                  db_path_entry.get(),
                                                                  password_entry.get()))
    generate_button.grid(row=6, column=0, columnspan=3)

    save_reports_button = tk.Button(root, text="Guardar Informes Individuales", bg=button_bg, fg=button_fg,
                                    command=lambda: threaded_function(save_individual_reports,
                                                                      start_date_entry.get(),
                                                                      end_date_entry.get(),
                                                                      db_path_entry.get(),
                                                                      password_entry.get()))
    save_reports_button.grid(row=7, column=0, columnspan=3)

    root.protocol("WM_DELETE_WINDOW", on_closing)    
    root.mainloop()

if __name__ == "__main__":
    create_gui()
